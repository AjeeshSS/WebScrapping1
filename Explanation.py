from bs4 import BeautifulSoup
import requests
import openpyxl 

# Creating a new Excel workbook
excel = openpyxl.Workbook()
# Printing the names of the sheets in the workbook
print(excel.sheetnames)

# Getting the active sheet in the workbook
sheet = excel.active
# Renaming the active sheet to 'Top Rated Movies'
sheet.title = 'Top Rated Movies'
# Printing the names of the sheets again to see the updated title
print(excel.sheetnames)

# Adding a header row to the sheet
sheet.append(['Movie name', 'Movie year', 'Movie rating'])


try:
    # Adding headers to make the request look like it's coming from a browser
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
    }
    
    # Sending a GET request to the IMDB Top 250 movies page with the custom headers
    source = requests.get('https://www.imdb.com/chart/top/', headers=headers)
    source.raise_for_status()  # Check if the request was successful (200 OK)
    
    # Parsing the HTML content of the page using BeautifulSoup
    soup = BeautifulSoup(source.text, 'html.parser')
    
    # Finding all 'li' tags with the specific class indicating movie items
    movie_items = soup.find_all('li', class_='ipc-metadata-list-summary-item')

    # Loop through each movie item to extract the desired information
    for item in movie_items:
        # Find the 'h3' tag with the class 'ipc-title__text' for the movie title
        title_tag = item.find('h3', class_='ipc-title__text')
        
        # Find the 'span' tag with the class 'cli-title-metadata-item' for the movie year
        year_tag = item.find('span', class_='cli-title-metadata-item')
        
        # Find the 'span' tag with the class 'ipc-rating-star--rating' for the movie rating
        rating_tag = item.find('span', class_='ipc-rating-star--rating')

        # Check if all tags were found to avoid errors when accessing their text
        if title_tag and year_tag and rating_tag:
            # Extract the text content of the tags
            title = title_tag.text
            year = year_tag.text
            rating = rating_tag.text
            
            # Print the extracted information in a formatted string
            print(f'Title: {title}, Year: {year}, Rating: {rating}')
            
            sheet.append([title, year, rating])

except Exception as e:
    # Print any exception that occurs during the request or parsing
    print(e)
    
excel.save('movie ratings.xlsx')
